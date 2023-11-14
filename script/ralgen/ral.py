# import binascii
# -*-coding:UTF-8-*-
# import copy
# from xlutils import copy
# from openpyxl import load_workbook
# from pandas
import xlrd
import os
import sys
import re
import time
# import argparse
from datetime import date, datetime

from openpyxl import workbook


def add_field_info(fp, attribute, subbits, field_name, default_val, descript):
    print("ral.py:add_filed_info")
    if field_name != '':
        # write field info
        num = '0123456789'
        val = ''
        hbit = ''
        lbit = ''
        for s in subbits:
            if s in num:
                val += s
            elif s == ':':
                hbit = val
                val = ''
        lbit = val
        if (hbit == '') and (lbit == ''):
            print('error: input field subbits of %s is []' % field_name)
            return
        elif hbit != '':
            width = int(hbit) - int(lbit) + 1
        else:
            width = 1

        if field_name == 'Reserved' :
            field_name += '_'+lbit
            attribute = 'ro'
        attribute = attribute.lower()
        if attribute == 'r':
            attribute = 'ro'
        elif attribute == 'w':
            attribute = 'wo'
        elif (attribute == '') or (attribute == None):
            print("Error: %s atrribute is None" % field_name)
            exit(1)
        if (field_name == 'Reserved') and (default_val == None or default_val == ''):
            default_val = width+'\'h0'
        str = """       field %s{
            bits %d;
            access %s;
            reset   %s;
        }\n""" % (field_name.upper(), width, attribute, default_val)
        # print(str)
        fp.writelines(str)


def gen_ralf(fp, sheet, merge):
    print("ral.py:gen_ralf")
    # print("********************************")
    # print(sheet.name)
    # print(sheet.nrows)
    # print(sheet.ncols)

    if sheet.cell_value(0, 0) != '':
        val = sheet.cell_value(0, 1)
    else:
        return
    val = sheet.cell_value(2, 0)
    if val == '':
        return
    start_index = 3
    next_index = 3
    # end_index = sheet.nrows
    end_index = sheet.merged_cells[-1][1]
    for index in merge:
        if index[0] > 2:
            next_index = index[0]
            if next_index == (start_index + 1):
                # create new block
                return
            elif next_index > (start_index + 1):
                # create register
                reg_addr = sheet.cell_value(start_index, 0)
                reg_name = sheet.cell_value(start_index, 1)
                str = " register %s @ 'h%s{\n" % (reg_name, reg_addr[2:])
                fp.write(str)
                str = "     left_to_right;\n"
                fp.write(str)
                for i in range(next_index - start_index):
                    attribute = sheet.cell_value(start_index+i, 2)
                    subbits = sheet.cell_value(start_index + i, 3)
                    field_name = sheet.cell_value(start_index + i, 4)
                    default_val = sheet.cell_value(start_index + i, 5)
                    descript = sheet.cell_value(start_index + i, 6)
                    add_field_info(fp, attribute, subbits, field_name, default_val, descript)
                start_index = next_index
                str = '}\n'
                fp.write(str)

    next_index = end_index
    if next_index > start_index + 1:
        # create register
        reg_addr = sheet.cell_value(start_index, 0)
        reg_name = sheet.cell_value(start_index, 1)
        str = " register %s @ 'h%s{\n" % (reg_name, reg_addr[2:])
        fp.write(str)
        for i in range(next_index - start_index):
            attribute = sheet.cell_value(start_index + i, 2)
            subbits = sheet.cell_value(start_index + i, 3)
            field_name = sheet.cell_value(start_index + i, 4)
            default_val = sheet.cell_value(start_index + i, 5)
            descript = sheet.cell_value(start_index + i, 6)
            add_field_info(fp, attribute, subbits, field_name, default_val, descript)
        start_index = next_index
        str = '}\n'
        fp.write(str)

    # add by wangxx
    # memory
    end_line = sheet.nrows
    # print(end_line)
    memory_line = sheet.merged_cells[-1][1]
    # print(memory_line)
    memory_info = sheet.cell_value(memory_line, 0)
    # print(memory_info)
    if memory_info != 'Memory Info':
        print("\nWarning: Please check the memory info!")
    memory_start = memory_line + 2

    # ========================================================================================================
    # add by wangxx(linux)
    # print(sheet.nrows)
    # print(sheet.ncols)
    # print(memory_start)
    # mem_name = sheet.cell_value(memory_start + 0, 0)
    # mem_start_addr = sheet.cell_value(memory_start + 0, 1)
    # mem_bits = sheet.cell_value(memory_start + 0, 2)
    # mem_size = sheet.cell_value(memory_start + 0, 3)
    # mem_attribute = sheet.cell_value(memory_start + 0, 4)
    # # print("****************************************")
    # print("memory_start: %s\n" % memory_start)
    # print("mem_name: %s\n" % mem_name)
    # print("mem_start_addr: %s\n" % mem_start_addr)
    # print("mem_bits: %s\n" % mem_bits)
    # print("mem_size: %s\n" % mem_size)
    # print("mem_attribute: %s\n" % mem_attribute)
    # if mem_name == '0' or mem_start_addr == '0' or mem_bits == '0' or mem_size == '0' or mem_attribute == '0':
    # print(sheet.cell_value(sheet.nrows - 1, 0))
    # print(sheet.cell_value(sheet.nrows - 1, 1))
    # print(sheet.cell_value(sheet.nrows - 1, 2))
    # print(sheet.cell_value(sheet.nrows - 1, 3))
    # print(sheet.cell_value(sheet.nrows - 1, 4))
    mem_name = sheet.cell_value(sheet.nrows - 1, 0)
    mem_start_addr = sheet.cell_value(sheet.nrows - 1, 1)
    mem_bits = sheet.cell_value(sheet.nrows - 1, 2)
    mem_size = sheet.cell_value(sheet.nrows - 1, 3)
    mem_attribute = sheet.cell_value(sheet.nrows - 1, 4)
    if mem_name == "name" or mem_start_addr == "start_addr" or mem_bits == "bits" or mem_size == "size" or mem_attribute == "Attribute":
        # RawVal = sheet.cell(memory_start, 0).value
        print("There is no mem info")
    else:
        # print(memory_start)
        if end_line > memory_start:
            for i in range(end_line - memory_start):
                print(i)
                mem_name = sheet.cell_value(memory_start + i, 0)
                mem_start_addr = sheet.cell_value(memory_start + i, 1)
                mem_bits = sheet.cell_value(memory_start + i, 2)
                mem_size = sheet.cell_value(memory_start + i, 3)
                mem_attribute = sheet.cell_value(memory_start + i, 4)
                str = " memory %s @ 'h%s {\n" % (mem_name, mem_start_addr[2:])
                fp.write(str)
                str = "     bits %s;\n" % int(mem_bits)
                str += "    size %s;\n" % mem_size
                str += "    access %s;\n" % mem_attribute.lower()
                str += "    }\n"
                # print(str)
                fp.write(str)

    # print(memory_start)
    # if end_line > memory_start:
    #     for i in range(end_line - memory_start):
    #         mem_name = sheet.cell_value(memory_start + i, 0)
    #         mem_start_addr = sheet.cell_value(memory_start + i, 1)
    #         mem_bits = sheet.cell_value(memory_start + i, 2)
    #         mem_size = sheet.cell_value(memory_start + i, 3)
    #         mem_attribute = sheet.cell_value(memory_start + i, 4)
    #         str = " memory %s @ 'h%s {\n" % (mem_name, mem_start_addr[2:])
    #         fp.write(str)
    #         str = "     bits %s;\n" % int(mem_bits)
    #         str += "    size %s;\n" % mem_size
    #         str += "    access %s;\n" % mem_attribute.lower()
    #         str += "    }\n"
    #         fp.write(str)


def read_excel(file_name):
    print("ral.py:read_excel")
    # workbook = xlrd.open_workbook(file_name, formatting_info=True)
    workbook = xlrd.open_workbook(file_name)
    # workbook = load_workbook(file_name)
    # book = xlrd.open_workbook(file_name)
    # workbook1 = copy.copy(book)
    # print("book is %s" % book)
    # print("workbook is %s" % workbook1)
    # print workbook.sheet_names()

    for sheet_name in workbook.sheet_names():
        sheet = workbook.sheet_by_name(sheet_name)
        # print sheet_name
        if sheet_name != 'Reserved':
            merge = []
            for (rlow, rhigh, clow, chigh) in sheet.merged_cells:
                merge.append([rlow, clow])
            merge.sort()
            ralf_name = "ral_%s.ralf" % sheet_name
            fp = open(ralf_name, 'w')
            str_block = "block %s_reg_top {\n" % sheet_name
            str_block += "      bytes 4;\n"
            fp.writelines(str_block)
            gen_ralf(fp, sheet, merge)
            str_block = "}\n"
            fp.write(str_block)
            print("Format conversion completed !!!\n")
            print("Program terminating ...")
            time.sleep(2)


def usage():
    print("ral.py:usage")
    msg = """%s: A register excel file convert to ralf file app.
    You should input excel file in command line.
    """

    msg_a = """positionla argument:
    excel_file The input excel file name, xls format only!
    """

    print(msg % os.path.basename(sys.argv[0]))
    print("Usage: %s excel_file" % os.path.basename(sys.argv[0]))
    print(msg_a)


if __name__ == '__main__':
    print("ral.py:__main__")
    usage()
    if len(sys.argv) == 1:
        file_name = input('Please input excel file name:\n')
    else:
        file_name = sys.argv[1]

    if file_name.strip() == '':
        print("\nError: No excel file name provided !!!")
        print("Program terminated ...")
        time.sleep(3)
        raise Exception("No excel file name provided !!!")
    # if os.path.exists(file_name) == False:
    if not os.path.exists(file_name):
        print("\nError: No excel file name provided !!!")
        print("Program terminated ...")
        time.sleep(3)
        raise Exception("No excel file name provided !!!")
    read_excel(file_name)

    print("there is need install sysnopsys VCS tools if you want translating \'xx.ralf\' into \'xx_regmodel.sv\'")
    # there is need install sysnopsys VCS tools if you want translating "xx.ralf" into "xx_regmodel.sv"
    # # workbook = xlrd.open_workbook(file_name, formatting_info=True)
    # workbook = xlrd.open_workbook(file_name)
    # for sheet_name in workbook.sheet_names():
    #     sheet = workbook.sheet_by_name(sheet_name)
    #     if sheet_name != 'Reserved':
    #         model_name = sheet_name
    #         # ralf2sv
    #         ralf2sv = 'ralgen -t %s_reg_top -l sv -uvm -a -c asF -B -full64 ral_%s.ralf -all_fields_rand -o%s_reg_top' % (model_name, model_name, model_name)
    #         os.system(ralf2sv)
